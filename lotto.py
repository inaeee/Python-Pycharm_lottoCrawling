import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

url="https://www.dhlottery.co.kr/gameResult.do?method=byWin&drwNo="

wb=Workbook()
ws=wb.active
ws.cell(1,1,"회차")
ws.cell(1,2,"당첨번호")
row=2

for i in range(1,878):
    data=requests.get(url+str(i))

    if data.status_code != requests.codes.ok:
        print("접속실패")
        exit()

    html=BeautifulSoup(data.text,"html.parser")
    lottos_containter=html.select(".win_result")
    #print(lottos_containter)
    date=html.select(".win_result > h4 > strong")
    number=html.select(".ball_645")
    numbers=number[0].text+","+number[1].text+","+number[2].text+","+number[3].text+","+number[4].text+","+number[5].text+","+number[6].text
    #print(date[0].text, numbers)

    ws.cell(row,1,date[0].text)
    ws.cell(row,2,numbers)
    row+=1

wb.save("lotto.xlsx")