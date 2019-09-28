import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

wb=Workbook()
ws=wb.active

ws.cell(1,1,"회차")
ws.cell(1,2,"숫자1")
ws.cell(1,3,"숫자2")
ws.cell(1,4,"숫자3")
ws.cell(1,5,"숫자4")
ws.cell(1,6,"숫자5")
ws.cell(1,7,"숫자6")
ws.cell(1,8,"보너스번호")

#for lotto in range(1,878):
for lotto in range(1, 10):
    #lotto=1
    row_num=lotto+1
    url=f"https://search.daum.net/nate?w=tot&rtmaxcoll=LOT&DA=LOT&q={lotto}회차%20로또"
    data=requests.get(url)

    if data.status_code != requests.codes.ok:
        print("접속실패")
        exit()

    html = BeautifulSoup(data.text, "html.parser")
    numbers=html.select("span.img_lotto")

    ws.cell(row_num,1,lotto)

    for index, number in enumerate(numbers,start=2):
        ws.cell(row_num,index,number.text)

wb.save("lotto_study.xlsx")